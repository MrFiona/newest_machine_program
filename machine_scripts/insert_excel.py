#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Time    : 2017-03-21 09:40
# Author  : MrFiona
# File    : insert_excel.py
# Software: PyCharm Community Edition

# TODO 2016-06-26 update: Add the keep_continuous parameters for the function that inserts the excel data
# TODO Description: keep_continuous：False --- Normal insertion of data  True --- Insert data from the specified location
# TODO 2016-06-27 update: 插入数据：目前采用位置标记法，根据位置变量控制而不是excel标记字符串位置定位
# TODO          1、有最新的周出现且需要覆盖  后续周位置向后移动n, n表示出现的新周数
# TODO          2、无最新的周出现且需要覆盖  保持既定位置不变
# TODO          3、正常情况不需要覆盖       保持既定位置不变

from __future__ import absolute_import

import copy
import os
import re
import sys
import time
import urllib2

import openpyxl
import xlsxwriter
import collections

reload(sys)
sys.setdefaultencoding('utf-8')
from machine_scripts import extract_NFV_data, extract_data
from machine_scripts.cache_mechanism import DiskCache
from machine_scripts.machine_config import MachineConfig
from setting_global_variable import SRC_EXCEL_DIR, PROGRAM_NAME_ID_DICT
from machine_scripts.public_use_function import (get_interface_config, judge_get_config)
from machine_scripts.common_interface_func import (verify_validity_url, hidden_data_by_column)
from machine_scripts.predict_extract_data import PredictGetData
from machine_scripts.common_interface_branch_func import analysis_url_address_string, traceback_print_info



class InsertDataIntoExcel(object):
    def __init__(self, silver_url_list, section_Silver_url_list, link_WW_week_string, verify_flag=False, purl_bak_string=None, cache=None,
                 logger=None, log_time=None, keep_continuous=False):
        logger.print_message('>>>>>>>>>> Please Wait ... The Program is being read for Excel Initialization <<<<<<<<<<', os.path.split(__file__)[1])
        self.cache = cache
        self.logger = logger
        self.log_time = log_time
        self.Silver_url_list = silver_url_list
        self.section_Silver_url_list = section_Silver_url_list
        # TODO 验证标志,默认不开启,即预测
        self.verify_flag = verify_flag
        self.purl_bak_string = purl_bak_string
        self.link_WW_week_string = link_WW_week_string
        # TODO 支持只覆盖某些周
        self.keep_continuous = keep_continuous
        self.date_string_list = []
        self.__file_name = os.path.split(__file__)[1]
        self.newest_week_type_string_list = []
        # TODO 控制预测周数据的获取渠道 True: 从接口渠道获取 False: 正常渠道获取
        self.predict_insert_flag = False

        # TODO 模板周数获取
        self.__WEEK_NUM = int(judge_get_config('week_num', self.purl_bak_string))

        # TODO 是否离线标记获取
        self.on_off_line_save_flag = judge_get_config('on_off_line_save_flag', self.purl_bak_string)

        # TODO 模板文件获取
        template_file = get_interface_config('template_file', self.purl_bak_string)
        self.logger.print_message('template_file:%s\t' % template_file, self.__file_name)

        # TODO 自定义选取周数 则判断此时实际最新周在全局周对应位置
        self.equal_silver_list_flag = False
        if self.keep_continuous == 'YES':
            # TODO 在选择是空模板的情况下
            self.actual_newest_week_position = self.Silver_url_list.index(self.section_Silver_url_list[0])
        # TODO 正常情况下取所有的最新周
        else:
            self.actual_newest_week_position = self.Silver_url_list.index(self.Silver_url_list[0])

        if self.Silver_url_list == self.section_Silver_url_list:
            self.equal_silver_list_flag = True

        if not os.path.exists(SRC_EXCEL_DIR):
            os.makedirs(SRC_EXCEL_DIR)

        self.workbook = xlsxwriter.Workbook(SRC_EXCEL_DIR + os.sep + self.purl_bak_string + '_%d_%s_%d_%s.xlsx'
                % (self.__WEEK_NUM, link_WW_week_string, len(self.Silver_url_list), self.log_time), options={'strings_to_urls': False})
        self.rb = openpyxl.load_workbook(template_file, data_only=False)

        # TODO 在excel工作簿中增加工作表单
        self.worksheet_save_miss = self.workbook.add_worksheet('Save-Miss')
        self.worksheet_mapping = self.workbook.add_worksheet('Mapping')
        self.worksheet_caseresult = self.workbook.add_worksheet('CaseResult')
        self.worksheet_test_time = self.workbook.add_worksheet('TestTime')
        self.worksheet_trend = self.workbook.add_worksheet('Trend')
        self.worksheet_newsi = self.workbook.add_worksheet('NewSi')
        self.worksheet_existing = self.workbook.add_worksheet('ExistingSi')
        self.worksheet_closesi = self.workbook.add_worksheet('ClosedSi')
        self.worksheet_rework = self.workbook.add_worksheet('Rework')
        self.worksheet_hw = self.workbook.add_worksheet('HW')
        self.worksheet_sw = self.workbook.add_worksheet('SW')
        self.worksheet_ifwi = self.workbook.add_worksheet('IFWI')
        self.worksheet_platform = self.workbook.add_worksheet('ValidationResult')
        self.worksheet_sw_original = self.workbook.add_worksheet('SW_Original')
        self.worksheet_ifwi_original = self.workbook.add_worksheet('IFWI_Original')
        self.worksheet_rework_info = self.workbook.add_worksheet('ReworkInfo')
        self.worksheet_hw_info = self.workbook.add_worksheet('HWInfo')
        self.worksheet_sw_info = self.workbook.add_worksheet('SWInfo')
        self.worksheet_ifwi_info = self.workbook.add_worksheet('IFWIInfo')
        self.worksheet_change = self.workbook.add_worksheet('Change History')

        # TODO 为excel对象设定显示格式
        self.yellow_data_format = self.workbook.add_format({'bg_color': '#FFFF66'})
        self.title_format = self.workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_size': 12})
        self.format1 = self.workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        self.url_format = self.workbook.add_format({'font_color': 'blue', 'underline': 1, 'align':'center', 'valign':'vcenter'})
        # TODO Add a format for the header cells.
        self.header_format = self.workbook.add_format({'border': 1, 'bg_color': '#C6EFCE', 'bold': True})
        self.red = self.workbook.add_format({'color': 'red'})
        # TODO 设置小数点格式部分
        self.reserve_two_decimal_format = self.workbook.add_format()
        self.reserve_zero_decimal_format = self.workbook.add_format()
        self.reserve_int_two_decimal_format = self.workbook.add_format()
        self.reserve_int_integer_format = self.workbook.add_format()
        self.reserve_int_integer_format.set_num_format( 0x01 )       #0
        self.reserve_int_two_decimal_format.set_num_format( 0x02 )   #0.00
        self.reserve_zero_decimal_format.set_num_format( 0x09 )      #0%
        self.reserve_two_decimal_format.set_num_format( 0x0a )       #0.00%

        # TODO 统一管理对象
        self.predict_extract_object = None
        self.save_miss_insert_bkc_string = 'default_bkc_string'
        self.predict_execute_flag = False
        self._predict_url = self.predict_week_insert()
        self.logger.print_message('_predict_url:\t%s' % self._predict_url, self.__file_name)

        self.logger.print_message('>>>>>>>>>> Excel Initialization End <<<<<<<<<<', self.__file_name)

    def return_predict_bkc_string(self):
        return self.save_miss_insert_bkc_string

    def predict_week_insert(self):
        program_id = PROGRAM_NAME_ID_DICT.get(self.purl_bak_string)
        if program_id is None:
            self.logger.print_message('The project id does not exist', self.__file_name, 50)
            return None
        else:
            self.logger.print_message('The project id exists:\t%s' % program_id, self.__file_name)
            predict_url = 'https://dcg-oss.intel.com/get_last_candidate_link/' + program_id
            try:
                return_result_url = urllib2.urlopen(predict_url).read()
                # return_result_url = 'https://dcg-oss.intel.com/test_report/test_report/6495/0/'
                # TODO 返回字符串不为0则未生成静态页面，数据从指定渠道获取
                if len(return_result_url) != 0:
                    self.predict_execute_flag = True
                    self.predict_insert_flag = True
                    self.logger.print_message('predict_insert_flag:\t%s' % self.predict_insert_flag, self.__file_name)
                    self.logger.print_message('return_result_url:\t%s' % return_result_url, self.__file_name)
                    # TODO 统一管理对象
                    self.predict_extract_object = PredictGetData(self.logger, return_result_url)
                    # TODO 插入save-miss表的bkc数据
                    self.save_miss_insert_bkc_string = self.predict_extract_object.return_save_miss_bkc_string()
                    return return_result_url
            except:
                # traceback_print_info(self.logger)
                self.logger.print_message('The Candidate date is not exists:\t%s' % program_id, self.__file_name, 50)

    def return_predict_execute_flag(self):
        return self.predict_execute_flag

    def get_url_list(self):
        return self.Silver_url_list

    def calculate_head_num(self, multiple, num, add_num=0, predict_calculate_flag=False):
        if not predict_calculate_flag:
            url_length = len(self.Silver_url_list)
            return multiple*(self.__WEEK_NUM - url_length + num) + add_num
        else:
            url_length = len(self.Silver_url_list) + 1
            return multiple*(self.__WEEK_NUM - url_length + num) + add_num

    def get_formula_data(self, data_type, wb_sheet_name):
        rb_sheet = self.rb.get_sheet_by_name(data_type)
        row, col = 1, 1
        for rs in rb_sheet.rows:
            col = 1
            for cell in rs:
                data = rb_sheet.cell(row=row, column=col).value
                if 'CaseResult' == data_type:
                    if row == 6 or (row in (2, 3) and ((col - 11) % 41) == 38) or (row == 2 and ((col - 11) % 41) == 40):
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_zero_decimal_format)
                    elif (row in (4, 5) and ((col - 11) % 41) in (6, 7, 8, 9, 38)):
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_integer_format)
                    else:
                        wb_sheet_name.write(row - 1, col - 1, data)

                elif 'Save-Miss' == data_type:
                    if 4 <= row <= 8:
                        if row != 7 and row != 6 and col == 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_two_decimal_format)
                        elif row in (6, 7) and col == 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_two_decimal_format)
                        elif row in (6, 7) and col != 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_integer_format)
                        else:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_zero_decimal_format)
                    elif 11 <= row <= 50 and col == 2:
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_two_decimal_format)
                    else:
                        wb_sheet_name.write(row - 1, col - 1, data)

                elif 'Trend' == data_type:
                    if (1 <= row <= len(self.Silver_url_list) + 1) and (8 <= col <= 10):
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_zero_decimal_format)
                    else:
                        wb_sheet_name.write(row - 1, col - 1, data)

                else:
                    wb_sheet_name.write(row - 1, col - 1, data)
                col += 1
            row += 1

    def insert_change_history_data(self):
        self.get_formula_data('Change History', self.worksheet_change)

    def insert_NewSi_data(self):
        self.logger.print_message('Start inserting [ New Sightings ] data.........', self.__file_name)
        #隐藏部分数据
        hidden_data_by_column(self.worksheet_newsi, self.Silver_url_list, 13, 1)
        # 获取公式并插入指定位置
        self.get_formula_data(u'NewSi', self.worksheet_newsi)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and  self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                #验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue

            self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

            if j == 0:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_new_sightings_data('New Sightings', self.verify_flag)
                if self.keep_continuous != 'YES':
                    self.newest_week_type_string_list.append(Silver_BkC_string)
                elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                    self.newest_week_type_string_list.append(Silver_BkC_string)
            else:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_new_sightings_data('New Sightings', True)
                if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                    self.newest_week_type_string_list.append(Silver_BkC_string)

            self.date_string_list.append(date_string)

            self.worksheet_newsi.conditional_format(4, self.calculate_head_num(13, j, 2), 250, self.calculate_head_num(13, j, 3),
                                                    {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
            self.worksheet_newsi.write(2, self.calculate_head_num(13, j, 1), Silver_BkC_string, self.format1)
            self.worksheet_newsi.write(2, self.calculate_head_num(13, j), date_string, self.format1)
            # self.worksheet_newsi.write(2, self.calculate_head_num(13, j, 12), 'Comments')

            if not effective_url_list and not header_list and not cell_data_list:
                continue

            header_list.append('comments')
            self.worksheet_newsi.write_row(2, self.calculate_head_num(13, j, 4), header_list, self.header_format)

            num_url_list = []
            for ele in effective_url_list:
                num_url_list.append(len(ele))

            for i in range(len(cell_data_list)):
                #插入非url部分数据
                self.worksheet_newsi.write_row(4 + i, self.calculate_head_num(13, j, 5), cell_data_list[i][1:-1])
                #插入url数据部分
                self.worksheet_newsi.write_url(4 + i, self.calculate_head_num(13, j, 4), effective_url_list[i][0], self.url_format, cell_data_list[i][0])
                if num_url_list[i] > 1:
                    self.worksheet_newsi.write_url(4 + i, self.calculate_head_num(13, j, 11), effective_url_list[i][1], self.url_format, cell_data_list[i][-1])
                elif num_url_list[i] == 1:
                    self.worksheet_newsi.write(4 + i, self.calculate_head_num(13, j, 11), cell_data_list[i][-1])

    def insert_ExistingSi_data(self):
        fail_url_list = []
        self.logger.print_message('Start inserting [ Existing Sightings ] data.........', self.__file_name)
        hidden_data_by_column(self.worksheet_existing, self.Silver_url_list, 13, 1)
        # 获取公式并插入指定位置
        self.get_formula_data('ExistingSi', self.worksheet_existing)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue

            self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

            if j == 0:
                Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_existing_sighting_data('Existing Sightings', self.verify_flag)
                if self.keep_continuous != 'YES':
                    self.newest_week_type_string_list.append(Silver_BkC_string)
                elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                    self.newest_week_type_string_list.append(Silver_BkC_string)
            else:
                Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_existing_sighting_data('Existing Sightings', True)
                if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                    self.newest_week_type_string_list.append(Silver_BkC_string)

            #增加一列comments
            header_list.append('comments')
            try:
                #标记True为红色
                self.worksheet_existing.conditional_format(4, self.calculate_head_num(13, j, 2), 250, self.calculate_head_num(13, j, 3),
                                                           {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_existing.write(2, self.calculate_head_num(13, j, 1), Silver_BkC_string, self.format1)
                self.worksheet_existing.write(2, self.calculate_head_num(13, j), date_string, self.format1)
                # self.worksheet_existing.write(2, self.calculate_head_num(13, j, 12), 'comment')

                if not url_list and not header_list and not cell_data_list:
                    continue

                # 写入表头行
                self.worksheet_existing.write_row(2, self.calculate_head_num(13, j, 4), header_list, self.header_format)

                # 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[7:])
                    line_num_list.append(num)

                nu = 4
                # 插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    if line_num_list[line] <= 1:
                        self.worksheet_existing.write_row(nu, self.calculate_head_num(13, j, 5), cell_data_list[line][1:7])
                        if line_num_list[line] == 0:
                            self.worksheet_existing.write_url(nu, self.calculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                            self.worksheet_existing.write(nu, self.calculate_head_num(13, j, 11), ' ')
                        elif line_num_list[line] == 1:
                            self.worksheet_existing.write_url(nu, self.calculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                            self.worksheet_existing.write_url(nu, self.calculate_head_num(13, j, 11), url_list[line][1], self.url_format, str(cell_data_list[line][7]))
                        nu += 1

                    elif line_num_list[line] >= 2:
                        length_merge = line_num_list[line]
                        self.worksheet_existing.write_url(nu, self.calculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                        self.worksheet_existing.write_row(nu, self.calculate_head_num(13, j, 5), cell_data_list[line][1:7],self.title_format)
                        for i in range(4, 11):
                            self.worksheet_existing.merge_range(nu, self.calculate_head_num(13, j, i),nu + line_num_list[line] - 1, self.calculate_head_num(13, j, i), '')
                        for m in range(length_merge):
                            self.worksheet_existing.write_url(nu + m, self.calculate_head_num(13, j, 11), url_list[line][m+1], self.url_format,str(cell_data_list[line][7 + m]))
                        nu += line_num_list[line]
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)
                fail_url_list.append(self.Silver_url_list[j])

    def insert_ClosedSi_data(self):
        self.logger.print_message('Start inserting [ Closed Sightings ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        fail_url_list = []
        self.get_formula_data('ClosedSi', self.worksheet_closesi)
        hidden_data_by_column(self.worksheet_closesi, self.Silver_url_list, 13, 1)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

                if j == 0:
                    Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_closed_sightings_data('Closed Sightings', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_closed_sightings_data('Closed Sightings', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                self.worksheet_closesi.conditional_format(4, self.calculate_head_num(13, j, 2), 250, self.calculate_head_num(13, j, 3),
                                                          {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_closesi.write(2, self.calculate_head_num(13, j, 1), Silver_BkC_string, self.format1)
                self.worksheet_closesi.write(2, self.calculate_head_num(13, j), date_string, self.format1)

                if not effective_url_list and not header_list and not cell_data_list:
                    continue

                self.worksheet_closesi.write_row(2, self.calculate_head_num(13, j, 4), header_list, self.header_format)

                num_url_list = []
                for ele in effective_url_list:
                    num_url_list.append(len(ele))

                for i in range(len(cell_data_list)):
                    #插入非url部分数据
                    self.worksheet_closesi.write_row(4 + i, self.calculate_head_num(13, j, 5), cell_data_list[i][1:-1])
                    #插入url数据部分
                    self.worksheet_closesi.write_url(4 + i, self.calculate_head_num(13, j, 4), effective_url_list[i][0], self.url_format, cell_data_list[i][0])
                    if num_url_list[i] > 1:
                        self.worksheet_closesi.write_url(4 + i, self.calculate_head_num(13, j, 11), effective_url_list[i][1], self.url_format, cell_data_list[i][-1])
                    elif num_url_list[i] == 1:
                        self.worksheet_closesi.write(4 + i, self.calculate_head_num(13, j, 11), cell_data_list[i][-1])
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)
                fail_url_list.append(self.Silver_url_list[j])

    def insert_Rework_data(self):
        self.logger.print_message('Start inserting [ HW Rework ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        self.get_formula_data('Rework', self.worksheet_rework)
        hidden_data_by_column(self.worksheet_rework, self.Silver_url_list, 3, 1)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)
                #最新的一周在验证标志未开启的情况下取Silver数据,否则取BKC数据
                if j == 0:
                    if self.purl_bak_string == 'Purley-FPGA' or self.purl_bak_string == 'NFVi':
                        Silver_BkC_string, object_string_list, date_string = obj.get_rework_data('HW Rework', self.verify_flag)
                    else:
                        Silver_BkC_string, object_string_list, date_string = obj.get_bak_rework_data('HW Rework', self.verify_flag)

                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    if self.purl_bak_string == 'Purley-FPGA' or self.purl_bak_string == 'NFVi':
                        Silver_BkC_string, object_string_list, date_string = obj.get_rework_data('HW Rework', True)
                    else:
                        Silver_BkC_string, object_string_list, date_string = obj.get_bak_rework_data('HW Rework', True)

                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                # 标记True为红色
                self.worksheet_rework.conditional_format(3, self.calculate_head_num(3, j), 300, self.calculate_head_num(3, j),
                                                         {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_rework.write(1, self.calculate_head_num(3, j, 1), date_string, self.format1)
                self.worksheet_rework.write(1, self.calculate_head_num(3, j, 2), Silver_BkC_string, self.format1)

                if not object_string_list:
                    continue

                if self.purl_bak_string == 'Purley-FPGA':
                    self.worksheet_rework.write_column(3, self.calculate_head_num(3, j, 1), object_string_list)
                else:
                    for row in range(len(object_string_list)):
                        if 0 <= row <= 9:
                            self.worksheet_rework.merge_range(row + 3, self.calculate_head_num(3, j, 1), row + 3, self.calculate_head_num(3, j, 2), '')
                            self.worksheet_rework.write(row + 3, self.calculate_head_num(3, j, 1), object_string_list[row])
                        else:
                            self.worksheet_rework.write_row(row + 3, self.calculate_head_num(3, j, 1), object_string_list[row])
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_ReworkInfo(self):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_rework_info, self.Silver_url_list, 1, 1)
        self.get_formula_data('ReworkInfo', self.worksheet_rework_info)

    def insert_HW_data(self):
        self.logger.print_message('Start inserting [ HW Configuration ] data.........', self.__file_name)

        # TODO FPGA 23周  DE 24周  最新格式适配边界位置
        if self.purl_bak_string == 'Purley-FPGA':
            FPGA_border_position = self.Silver_url_list.index('https://dcg-oss.intel.com/ossreport/auto/Purley-FPGA/Silver/2017%20WW23/6131_Silver.html')
            self.logger.print_message('FPGA_border_position:\t%s' % FPGA_border_position, self.__file_name)
        elif self.purl_bak_string == 'Bakerville':
            DE_border_position = self.Silver_url_list.index('https://dcg-oss.intel.com/ossreport/auto/Bakerville/Silver/2017%20WW24/6170_Silver.html')
            self.logger.print_message('DE_border_position:\t%s' % DE_border_position, self.__file_name)

        # 获取公式并插入指定位置
        self.get_formula_data('HW', self.worksheet_hw)
        hidden_data_by_column(self.worksheet_hw, self.Silver_url_list, 18, 1)

        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue

            self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)

            # TODO 分项目获取数据 2017-06-09
            if self.purl_bak_string == 'NFVi':
                obj = extract_NFV_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)
            else:
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

            if j == 0:
                if self.purl_bak_string == 'Purley-FPGA' or self.purl_bak_string == 'NFVi':
                    if self.purl_bak_string == 'Purley-FPGA' and FPGA_border_position >= j:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_lastest_FPGA_hw_data('HW Configuration', self.verify_flag)
                    else:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_hw_data('HW Configuration', self.verify_flag)
                else:
                    if DE_border_position >= j:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_lastest_bak_hw_data('HW Configuration', self.verify_flag)
                        if Silver_BkC_string == 'Error' and date_string == 'Error' and not row_coordinates_list and not column_coordinates_list and not cell_data_list:
                            Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_bak_hw_data('HW Configuration', self.verify_flag)
                    else:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_bak_hw_data('HW Configuration', self.verify_flag)

                if self.keep_continuous != 'YES':
                    self.newest_week_type_string_list.append(Silver_BkC_string)
                elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                    self.newest_week_type_string_list.append(Silver_BkC_string)

            else:
                if self.purl_bak_string == 'Purley-FPGA' or self.purl_bak_string == 'NFVi':
                    if self.purl_bak_string == 'Purley-FPGA' and FPGA_border_position >= j:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_lastest_FPGA_hw_data('HW Configuration', True)
                    else:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_hw_data('HW Configuration', True)
                # TODO DE
                else:
                    if DE_border_position >= j:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_lastest_bak_hw_data('HW Configuration', True)
                        if Silver_BkC_string == 'Error' and date_string == 'Error' and not row_coordinates_list and not column_coordinates_list and not cell_data_list:
                            Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_bak_hw_data('HW Configuration', True)
                    else:
                        Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_bak_hw_data('HW Configuration', True)

                if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                    self.newest_week_type_string_list.append(Silver_BkC_string)
            # 合并单元格
            self.worksheet_hw.merge_range(3, self.calculate_head_num(18, j), 9, self.calculate_head_num(18, j), "", self.title_format)
            self.worksheet_hw.write_rich_string(3, self.calculate_head_num(18, j), self.red, date_string, self.title_format)

            self.worksheet_hw.write(2, self.calculate_head_num(18, j), Silver_BkC_string, self.format1)
            # 标记True为红色
            self.worksheet_hw.conditional_format(2, self.calculate_head_num(18, j, 1), 10, self.calculate_head_num(13, j, 17),
                                                 {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})

            #判断这周有无数据， 例如：第40周的数据为空
            if not row_coordinates_list and not column_coordinates_list and not cell_data_list:
                continue

            try:
                # 写入工作表的第二行,现在只支持system 1 到HWsystem 15，截取15行数据
                # self.worksheet_hw.write_row(1, self.calculate_head_num(18, j, 2), row_coordinates_list[:15])
                #按行插入数据
                self.worksheet_hw.write_column(3, self.calculate_head_num(18, j, 2), column_coordinates_list)
                for row_num in range(len(cell_data_list)):
                    #每个元素限制为10列 [:15]
                    self.worksheet_hw.write_row(3 + row_num, self.calculate_head_num(18, j, 3), cell_data_list[row_num][:15])
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_HWInfo(self):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_hw_info, self.Silver_url_list, 1, 1)
        self.get_formula_data('HWInfo', self.worksheet_hw_info)

    def insert_SW_Original_data(self):
        self.logger.print_message('Start inserting [ SW_Original Configuration ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw_original, self.Silver_url_list, 9, 1)
        self.get_formula_data('SW_Original', self.worksheet_sw_original)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

                if j == 0:
                    if 'Purley-FPGA/Silver/2017%20WW31' in self.Silver_url_list[j]:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data_1('SW Configuration', self.verify_flag)
                    else:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data('SW Configuration', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    if 'Purley-FPGA/Silver/2017%20WW31' in self.Silver_url_list[j]:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data_1('SW Configuration', True)
                    else:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data('SW Configuration', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                self.worksheet_sw_original.write_rich_string(1, self.calculate_head_num(9, j, 1), self.red, date_string, self.header_format)
                self.worksheet_sw_original.write_rich_string(1, self.calculate_head_num(9, j, 2), self.red, Silver_BkC_string, self.header_format)

                if not header_list and not cell_data_list:
                    continue

                # 标记True为红色
                self.worksheet_sw_original.conditional_format(3, self.calculate_head_num(9, j), 40, self.calculate_head_num(9, j, 1),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_sw_original.conditional_format(3, self.calculate_head_num(9, j, 2), 40, self.calculate_head_num(9, j, 2),
                                                             {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_sw_original.write_row(2, self.calculate_head_num(9, j, 5), header_list, self.header_format)
                #插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)
                nu = 3
                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                #插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw_original.write(nu, self.calculate_head_num(9, j, 5), first_insert_data[line])
                        self.worksheet_sw_original.write_url(nu, self.calculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_original.write(nu, self.calculate_head_num(9, j, 7), second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw_original.write_url(nu, self.calculate_head_num(9, j, 8), url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw_original.write(nu, self.calculate_head_num(9, j, 8), cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw_original.merge_range(nu, self.calculate_head_num(9, j, 5), nu + line_num_list[line] -1, self.calculate_head_num(9, j, 5), first_insert_data[line], self.title_format)
                        self.worksheet_sw_original.merge_range(nu, self.calculate_head_num(9, j, 6), nu + line_num_list[line] - 1, self.calculate_head_num(9, j, 6), '')
                        self.worksheet_sw_original.write_url(nu, self.calculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_original.merge_range(nu, self.calculate_head_num(9, j, 7), nu + line_num_list[line] -1, self.calculate_head_num(9, j, 7), second_insert_data[line], self.title_format)

                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw_original.write_url(i, self.calculate_head_num(9, j, 8), url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw_original.write(line_num_list[line] + nu - 1, self.calculate_head_num(9, j, 8), cell_data_list[line][3:][-1])
                        nu += line_num_list[line]

            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_SW_data(self):
        self.logger.print_message('Start inserting [ SW Configuration ] data.........', self.__file_name)
        #获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw, self.Silver_url_list, 9, 1)
        self.get_formula_data('SW', self.worksheet_sw)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

                if j == 0:
                    if 'Purley-FPGA/Silver/2017%20WW31' in self.Silver_url_list[j]:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data_1('SW Configuration', self.verify_flag)
                    else:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data('SW Configuration', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    if 'Purley-FPGA/Silver/2017%20WW31' in self.Silver_url_list[j]:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data_1('SW Configuration', True)
                    else:
                        Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list = obj.get_sw_data('SW Configuration', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                self.worksheet_sw.write_rich_string(1, self.calculate_head_num(9, j, 1), self.red, date_string, self.header_format)
                self.worksheet_sw.write_rich_string(1, self.calculate_head_num(9, j, 2), self.red, Silver_BkC_string, self.header_format)

                if not header_list and not cell_data_list:
                    continue

                # Set up some formats to use.
                self.worksheet_sw.conditional_format(3, self.calculate_head_num(9, j), 40, self.calculate_head_num(9, j, 1),
                                                     {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_sw.conditional_format(3, self.calculate_head_num(9, j, 2), 40, self.calculate_head_num(9, j, 2),
                                                     {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_sw.write_row(2, self.calculate_head_num(9, j, 5), header_list, self.header_format)
                # 插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)

                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                #处理并插入数据
                insert_data = zip( cell_data_list, url_list, line_num_list )
                insert_data.sort(key=lambda x: x[0][0].upper())

                cell_data_list, url_list, line_num_list = zip( *insert_data )
                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                # 插入数据第一列到第三列
                nu = 3
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw.write(nu, self.calculate_head_num(9, j, 5), first_insert_data[line])
                        self.worksheet_sw.write_url(nu, self.calculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.write(nu, self.calculate_head_num(9, j, 7), second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw.write_url(nu, self.calculate_head_num(9, j, 8), url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw.write(nu, self.calculate_head_num(9, j, 8), cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw.merge_range(nu, self.calculate_head_num(9, j, 5),
                                                      nu + line_num_list[line] - 1, self.calculate_head_num(9, j, 5), first_insert_data[line], self.title_format)
                        self.worksheet_sw.merge_range(nu, self.calculate_head_num(9, j, 6),
                                                      nu + line_num_list[line] - 1, self.calculate_head_num(9, j, 6), '')
                        self.worksheet_sw.write_url(nu, self.calculate_head_num(9, j, 6),
                                                    url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.merge_range(nu, self.calculate_head_num(9, j, 7),
                                                      nu + line_num_list[line] - 1, self.calculate_head_num(9, j, 7), second_insert_data[line], self.title_format)
                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw.write_url(i, self.calculate_head_num(9, j, 8),
                                                        url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw.write(line_num_list[line] + nu - 1, self.calculate_head_num(9, j, 8), cell_data_list[line][3:][-1])
                        nu += line_num_list[line]

            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_SWInfo(self):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw_info, self.Silver_url_list, 1, 1)
        self.get_formula_data('SWInfo', self.worksheet_sw_info)

    def insert_IFWI_Original_data(self):
        self.logger.print_message('Start inserting [ IFWI_Original Configuration ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi_original, self.Silver_url_list, 6, 1)
        self.get_formula_data('IFWI_Original', self.worksheet_ifwi_original)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)
                if j == 0:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                # Set up some formats to use.
                self.worksheet_ifwi_original.conditional_format(4, self.calculate_head_num(6, j), 50, self.calculate_head_num(6, j, 1),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_ifwi_original.conditional_format(4, self.calculate_head_num(6, j, 2), 49, self.calculate_head_num(6, j, 2),
                                                             {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_ifwi_original.write_rich_string(1, self.calculate_head_num(6, j, 1), self.red, date_string, self.header_format)
                self.worksheet_ifwi_original.write_rich_string(1, self.calculate_head_num(6, j, 2), self.red, Silver_BkC_string, self.format1)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_ifwi_original.write_row(3, self.calculate_head_num(6, j, 3), header_list, self.header_format)
                #插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    #以数字开头的元素前面加Nic
                    match_obj = re.match('\s+\d+', str(cell_data_list[ele][0]))
                    if match_obj:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                for begin in range(len(cell_data_list)):
                    self.worksheet_ifwi_original.write_row(begin + 4,  self.calculate_head_num(6, j, 3), cell_data_list[begin])
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_IFWI_data(self):
        self.logger.print_message('Start inserting [ IFWI Configuration ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi, self.Silver_url_list, 6, 1)
        self.get_formula_data('IFWI', self.worksheet_ifwi)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)
                if j == 0:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration',self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                self.worksheet_ifwi.write_rich_string(1, self.calculate_head_num(6, j, 1), self.red, date_string, self.header_format)
                self.worksheet_ifwi.write_rich_string(1, self.calculate_head_num(6, j, 2), self.red, Silver_BkC_string, self.format1)

                if not header_list and not cell_data_list:
                    continue

                # Set up some formats to use.
                self.worksheet_ifwi.conditional_format(4, self.calculate_head_num(6, j), 50, self.calculate_head_num(6, j, 1),
                                                       {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_ifwi.conditional_format(4, self.calculate_head_num(6, j, 2), 49, self.calculate_head_num(6, j, 2),
                                                       {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_ifwi.write_row(4, self.calculate_head_num(6, j, 3), header_list, self.header_format)
                #插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    #以数字开头的元素前面加Nic
                    match_obj = re.match('^\d+', str(cell_data_list[ele][0]))
                    if match_obj:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                cell_data_list.sort(key=lambda x: x[0].upper())
                for begin in range(len(cell_data_list)):
                    self.worksheet_ifwi.write_row(begin + 4,  self.calculate_head_num(6, j, 3), cell_data_list[begin])
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_IFWIInfo(self):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi_info, self.Silver_url_list, 1, 1)
        self.get_formula_data('IFWIInfo', self.worksheet_ifwi_info)

    def insert_Platform_data(self):
        self.logger.print_message('Start inserting [ Platform Integration Validation Result ] data.........', self.__file_name)
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_platform, self.Silver_url_list, 12, 1)
        self.get_formula_data('ValidationResult', self.worksheet_platform)
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

                if j == 0:
                    Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_platform_data('Platform Integration Validation Result', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_BkC_string)
                else:
                    Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_platform_data('Platform Integration Validation Result', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_BkC_string)

                self.worksheet_platform.write_rich_string(1, self.calculate_head_num(12, j, 1), self.red, Silver_BkC_string, self.format1)
                self.worksheet_platform.write_rich_string(2, self.calculate_head_num(12, j), self.red, date_string, self.format1)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_platform.write_row(2, self.calculate_head_num(12, j, 1), header_list, self.header_format)
                # 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[10:])
                    line_num_list.append(num)
                #有可能出现多个多行的情况，自适应合并列的变化
                merge_width = len(cell_data_list) + 2
                for k in line_num_list:
                    if k > 1:
                        merge_width += k - 1

                self.worksheet_platform.merge_range(2, self.calculate_head_num(12, j), merge_width, self.calculate_head_num(12, j), '', self.title_format)
                self.worksheet_platform.write_rich_string(2, self.calculate_head_num(12, j), self.red, date_string, self.title_format)
                # 标记True为红色
                self.worksheet_platform.conditional_format(3, self.calculate_head_num(12, j, 5), len(cell_data_list) + 3, self.calculate_head_num(12, j, 5),
                                                           {'type': 'cell', 'criteria': '>', 'value': 0, 'format': self.format1})
                nu = 3
                # 插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    #将数字字符转化为数字
                    temp_list = copy.deepcopy(cell_data_list[line])
                    for ele in range(len(temp_list)):
                        if temp_list[ele].isdigit():
                            temp_list[ele] = int(temp_list[ele])

                    if line_num_list[line] == 1:
                        self.worksheet_platform.write_row(nu, self.calculate_head_num(12, j, 1), temp_list)
                        if url_list[line]:
                            self.worksheet_platform.write_url(nu, self.calculate_head_num(12, j, 11), url_list[line][0], self.url_format, str(temp_list[10]))
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = line_num_list[line]
                        for i in range(1, 11):
                            self.worksheet_platform.merge_range(nu, self.calculate_head_num(12, j, i),nu + line_num_list[line] - 1,self.calculate_head_num(12, j, i), '')
                        self.worksheet_platform.write_row(nu, self.calculate_head_num(12, j, 1), temp_list[:10], self.title_format)
                        for m in range(length_merge):
                            self.worksheet_platform.write_url(nu+m, self.calculate_head_num(12, j, 11), url_list[line][m], self.url_format, str(temp_list[10+m]))
                        nu += line_num_list[line]
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_Mapping(self):
        Silver_url_list = self.Silver_url_list
        # TODO 当自定义覆盖相应周数据时 需要处理日期对齐 start
        if self.keep_continuous == 'YES':
            Silver_url_list = self.section_Silver_url_list

        all_insert_index_dict = collections.OrderedDict()
        temp_insert_index_dict = collections.OrderedDict()
        for index, value in enumerate(self.Silver_url_list):
            all_insert_index_dict[value] = index + self.__WEEK_NUM + 6 - len(self.Silver_url_list) - 1

        for value in Silver_url_list:
            if value in all_insert_index_dict:
                url_sep_list = value.split('/')
                week_info = url_sep_list[-2]
                effective_week_info = week_info.replace('%20', '')
                temp_insert_index_dict[effective_week_info] = all_insert_index_dict[value]
        # TODO 当自定义覆盖相应周数据时 需要处理日期对齐 end

        height = self.__WEEK_NUM - len(self.Silver_url_list) + 5
        for row in range(height - 5 - 1):
            self.worksheet_mapping.set_row(row=row + 6, options={'hidden':True})

        self.worksheet_mapping.set_column(5, height - 1 - 1, options={'hidden': True})

        # 获取公式并插入指定位置
        self.get_formula_data('Mapping', self.worksheet_mapping)
        # 标记True为红色
        self.worksheet_mapping.conditional_format(self.__WEEK_NUM + 7, 5, 250, self.__WEEK_NUM + 6, {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
        self.worksheet_mapping.conditional_format(self.__WEEK_NUM + 7, self.__WEEK_NUM + 7, 250, self.__WEEK_NUM + 41, {'type': 'cell', 'criteria': '>', 'value': 0, 'format': self.format1})
        # TODO 解析url
        insert_date_list = analysis_url_address_string(Silver_url_list)
        # TODO 处理日期对齐
        if self.keep_continuous == 'YES':
            for value in temp_insert_index_dict:
                self.worksheet_mapping.write(self.__WEEK_NUM + 6, temp_insert_index_dict[value], value, self.header_format)
                self.worksheet_mapping.write(temp_insert_index_dict[value] + 1, self.__WEEK_NUM + 6, value, self.header_format)
        else:
            self.worksheet_mapping.write_row(self.__WEEK_NUM + 6, self.__WEEK_NUM + 6 - len(Silver_url_list) - 1, insert_date_list, self.header_format)
            self.worksheet_mapping.write_column(self.__WEEK_NUM + 6 - len(Silver_url_list) - 1 + 1, self.__WEEK_NUM + 6, insert_date_list, self.header_format)

    def insert_CaseResult(self):
        self.logger.print_message('Start inserting [ CaseResult ] data.........', self.__file_name)
        self.worksheet_caseresult.set_column(firstcol=11, lastcol=41 * (self.__WEEK_NUM - len(self.Silver_url_list) - 1) + 10, options={'hidden': True})

        # 获取公式并插入指定位置
        self.get_formula_data('CaseResult', self.worksheet_caseresult)
        yellow_header_format = self.workbook.add_format({'bg_color': '#FFFF66'})
        brown_header_format = self.workbook.add_format({'bg_color': '#FF9933'})
        dark_red_header_format = self.workbook.add_format({'bg_color': '#FF2D2D'})
        green_header_format = self.workbook.add_format({'bg_color': '#79FF79'})
        blue_header_format = self.workbook.add_format({'bg_color': '#97CBFF'})
        caseresult_header_format = self.workbook.add_format({'border': 1, 'align': 'center', 'bg_color': '#97CBFF', 'bold': 1, 'font_size': 13})
        caseresult_last_header_format = self.workbook.add_format({'bold': 1, 'align': 'center', 'font_size': 13})
        start_string = ['START']*31
        yellow_promote_string = ['Reported FAILED Sightings in test result', 'Reported Sightings in test result',
                                 'Missed Key Sightings', "Missed Key Sightings which is not due to new added test case.  "
                                "If this issue can be found by other selected test case, consider it as not missed.  "
                                "If this issue failed one test case, and blocked some other test cases, then those blocked test case don't have to be selected.",
                                 'Missed Sightings in test plan comparing to test result?', 'Missed Test Case in Test Plan?',
                                 'New added Test Case comparing to Test Plan Pool?', 'Different Test Case between Test Plan and test result?',
                                 'New added Test Case comparing to entire Test Pool?']
        brown_promote_string = ['Removed Test Case in Test Result?', 'Different Test Case between Test Plan Pool and previous Test Result?',
                                'New Sighting with old test case in test result?', 'Missed New Sighting with old test case in test plan?',
                                'ExistingSighting in previous test result?', 'Fixed Sighting in release notes?',
                                'Fixed Sighting in test result?', 'Missed Fixed Sighting in test plan?',
                                'New Add Test Case comparing to previous test plan pool?', 'Mapped in Mapping table?',
                                'NOT Found in Mapping table?', 'Covered new Sighting?', 'Covered existing sighting?',
                                'Covered Sighting?', 'Missed Sighting in test plan', 'No run in the past x(x<=10) weeks?',
                                'No Issue found in the latest y(y<=10) tests?', 'Basic Case?', 'Selected override',
                                'Efforts', 'Not in actual test result but is selected in Test Plan.', 'Selected(z = threshold value)?']
        last_header_list = ['Domain', 'Category', 'Case', 'New Added Test Case Comparing to Previous Test Case Pool',
                            'New Added Test Case Comparing to Entire Test Case Pool']
        for j in range(len(self.Silver_url_list)):
            # TODO 不在url列表范围则跳过不覆盖
            if self.keep_continuous == 'YES' and self.Silver_url_list[j] not in self.section_Silver_url_list:
                continue
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self.Silver_url_list[j], self.logger)
                if not is_validity:
                    continue
            try:
                self.logger.print_message('Start inserting the corresponding data for url [ %s ]' % self.Silver_url_list[j], self.__file_name)
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache, insert_flag=True, logger=self.logger, purl_bak_string=self.purl_bak_string)

                if j == 0:
                    date_string, Silver_Gold_BKC_string, tip_string, header_list, cell_data_list = obj.get_caseresult_data('Platform Integration Validation Result', self.verify_flag)
                    if self.keep_continuous != 'YES':
                        self.newest_week_type_string_list.append(Silver_Gold_BKC_string)
                    elif self.keep_continuous == 'YES' and self.equal_silver_list_flag:
                        self.newest_week_type_string_list.append(Silver_Gold_BKC_string)
                else:
                    date_string, Silver_Gold_BKC_string, tip_string, header_list, cell_data_list = obj.get_caseresult_data('Platform Integration Validation Result', True)
                    if self.keep_continuous == 'YES' and j == self.actual_newest_week_position:
                        self.newest_week_type_string_list.append(Silver_Gold_BKC_string)

                #按照格式写入一些数据
                self.worksheet_caseresult.write(6, self.calculate_head_num(41, j, 1+11+1), Silver_Gold_BKC_string, self.format1)

                self.worksheet_caseresult.write(2, self.calculate_head_num(41, j, 11), date_string, self.format1)

                if not header_list and not cell_data_list:
                    continue

                for i in range(len(yellow_promote_string)):
                    self.worksheet_caseresult.write_rich_string(6, self.calculate_head_num(41, j, 5+i+11), yellow_promote_string[i], yellow_header_format)
                    self.worksheet_caseresult.write_rich_string(7, self.calculate_head_num(41, j, 5+i+11), start_string[:9][i], yellow_header_format)
                for i in range(len(brown_promote_string)):
                    self.worksheet_caseresult.write_rich_string(6, self.calculate_head_num(41, j, 14+i+11), brown_promote_string[i], brown_header_format)
                self.worksheet_caseresult.write_row(7, self.calculate_head_num(41, j, 14+11), start_string[9:])

                self.worksheet_caseresult.write_row(7, self.calculate_head_num(41, j, 11), header_list, caseresult_header_format)
                self.worksheet_caseresult.write_row(7, self.calculate_head_num(41, j, 36+11), last_header_list, caseresult_last_header_format)

                # 标记True为红色
                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 3+11), 250, self.calculate_head_num(41, j, 3+11),
                                                             {'type': 'text', 'criteria': 'containing', 'value': 'FAILED', 'format': dark_red_header_format})
                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 3+11), 250, self.calculate_head_num(41, j, 3+11),
                                                             {'type': 'text', 'criteria': 'containing', 'value': 'PASSED', 'format': green_header_format})
                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 3 + 11), 250, self.calculate_head_num(41, j, 3 + 11),
                                                             {'type': 'text', 'criteria': 'containing', 'value': 'BLOCKED', 'format': brown_header_format})
                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 13+11), 250, self.calculate_head_num(41, j, 13+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': False, 'format': blue_header_format})
                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 23+11), 250, self.calculate_head_num(41, j, 23+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.format1})
                for line in range(len(cell_data_list)):
                    self.worksheet_caseresult.write_row(8 + line, self.calculate_head_num(41, j, 11), cell_data_list[line])

                self.worksheet_caseresult.conditional_format(8, self.calculate_head_num(41, j, 9+11), 250, self.calculate_head_num(41, j, 35+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
            except:
                self.logger.print_message('The data crawling failed for the %dth url [ %s ]' % (j + 1, self.Silver_url_list[j]), self.__file_name)

    def insert_save_miss_data(self):
        Silver_url_list = self.Silver_url_list
        # TODO 当自定义覆盖相应周数据时 需要处理日期对齐 start
        if self.keep_continuous == 'YES':
            Silver_url_list = self.section_Silver_url_list

        all_insert_index_dict = collections.OrderedDict()
        temp_insert_index_dict = collections.OrderedDict()
        for index, value in enumerate(self.Silver_url_list):
            all_insert_index_dict[value] = index + self.__WEEK_NUM + 2 - len(self.Silver_url_list)

        for value in Silver_url_list:
            if value in all_insert_index_dict:
                url_sep_list = value.split('/')
                week_info = url_sep_list[-2]
                effective_week_info = week_info.replace('%20', '')
                temp_insert_index_dict[effective_week_info] = all_insert_index_dict[value]
        # print 'temp_insert_index_dict:\t', temp_insert_index_dict
        # TODO 当自定义覆盖相应周数据时 需要处理日期对齐 end
        self.worksheet_save_miss.set_column(firstcol=2, lastcol=self.__WEEK_NUM - len(self.Silver_url_list) -1 + 1, options={'hidden': True})
        # 获取公式并插入指定位置
        yellow_header_format = self.workbook.add_format({'bg_color': '#FFFF66'})
        first_string_list = ["Save Test Case based on this week's Test case pool", "Miss Sightings in test plan comparing to test result (%)?",
                             "Miss Sightings in test plan comparing to test result?", "Total Sighting", "Save Effort"]
        self.get_formula_data('Save-Miss', self.worksheet_save_miss)
        # TODO 解析url
        insert_date_list = analysis_url_address_string(Silver_url_list)
        # TODO 处理日期对齐
        if self.keep_continuous == 'YES':
            for value in temp_insert_index_dict:
                self.worksheet_save_miss.write(2, temp_insert_index_dict[value], value, self.format1)
        else:
            self.worksheet_save_miss.write_row(2, self.__WEEK_NUM + 2 - len(Silver_url_list), insert_date_list, self.format1)

        for i in range(len(first_string_list)):
            self.worksheet_save_miss.write_rich_string(3 + i, 0, first_string_list[i], yellow_header_format)

        length_week = len(Silver_url_list)

        self.worksheet_save_miss.conditional_format(35, 0, 46, 102, {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(3, 1, 3, 1, {'type': 'cell', 'criteria': '<', 'value': 0.20, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(4, 102 - length_week, 4, 101, {'type': 'cell', 'criteria': '>', 'value': 0.005, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(5, 102 - length_week, 5, 101, {'type': 'cell', 'criteria': '>', 'value': 0.5, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(10, 102 - length_week, 12, 101, {'type': 'cell', 'criteria': '<', 'value': 0.5, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(14, 102 - length_week, 21, 101, {'type': 'cell', 'criteria': '<', 'value': 0.5, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(24, 102 - length_week, 32, 101, {'type': 'cell', 'criteria': '>', 'value': 0.5, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(35, 102 - length_week, 45, 101, {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
        self.worksheet_save_miss.conditional_format(48, 102 - length_week, 58, 101, {'type': 'cell', 'criteria': '<', 'value': 0.5, 'format': self.format1})

        self.worksheet_save_miss.conditional_format(13, 102 - length_week, 13, 101, {'type': 'cell', 'criteria': 'between', 'minimum': -0.5, 'maximum': 0.5, 'format': self.format1})

    def insert_trend_data(self):
        self.get_formula_data('Trend', self.worksheet_trend)

    def insert_test_time_data(self):
        # 获取公式并插入指定位置
        self.get_formula_data('TestTime', self.worksheet_test_time)

    def return_name(self):
        return self.__class__.__dict__

    def close_workbook(self):
        self.workbook.close()

    def return_newest_week_type_string_list(self):
        return self.newest_week_type_string_list

    def predict_insert_NewSi_data(self):
        predict_insert_location = self.calculate_head_num(13, 0, 0, True)
        self.worksheet_newsi.write(2, predict_insert_location + 1, 'Candidate', self.format1)
        self.worksheet_newsi.write(2, predict_insert_location, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_ExistingSi_data(self):
        predict_insert_location = self.calculate_head_num(13, 0, 0, True)
        self.worksheet_existing.write(2, predict_insert_location + 1, 'Candidate', self.header_format)
        self.worksheet_existing.write(2, predict_insert_location, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_ClosedSi_data(self):
        predict_insert_location = self.calculate_head_num(13, 0, 0, True)
        self.worksheet_closesi.write(2, predict_insert_location + 1, 'Candidate', self.format1)
        self.worksheet_closesi.write(2, predict_insert_location, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_Rework_data(self):
        predict_insert_location = self.calculate_head_num(3, 0, 0, True)
        self.worksheet_rework.write(1, predict_insert_location + 2, 'Candidate', self.format1)
        self.worksheet_rework.write(1, predict_insert_location + 1, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_HW_data(self):
        predict_insert_location = self.calculate_head_num(18, 0, 0, True)
        self.worksheet_hw.write(2, predict_insert_location, 'Candidate', self.format1)
        self.worksheet_hw.merge_range(3, predict_insert_location, 9, predict_insert_location, "", self.title_format)
        self.worksheet_hw.write_rich_string(3, predict_insert_location, self.red, self.save_miss_insert_bkc_string, self.title_format)

    def predict_insert_SW_data(self):
        if self.predict_insert_flag and self._predict_url:
            predict_insert_location = self.calculate_head_num(9, 0, 0, True)
            self.logger.print_message('SW predict_insert_location:\t%s' % predict_insert_location, self.__file_name)
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self._predict_url, self.logger)
                if not is_validity:
                    return
            try:
                Silver_BkC_string, header_length, url_list, header_list, cell_data_list = self.predict_extract_object.predict_get_sw_data()
                self.worksheet_sw.write_rich_string(1, predict_insert_location + 1, self.red, self.save_miss_insert_bkc_string, self.header_format)
                self.worksheet_sw.write_rich_string(1, predict_insert_location + 2, self.red, Silver_BkC_string, self.header_format)

                if not header_list and not cell_data_list:
                    return

                # Set up some formats to use.
                self.worksheet_sw.conditional_format(3, predict_insert_location, 40, predict_insert_location + 1,
                                                     {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_sw.conditional_format(3, predict_insert_location + 2, 40, predict_insert_location + 2,
                                                     {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_sw.write_row(2, predict_insert_location + 5, header_list, self.header_format)
                # 插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)

                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                # 处理并插入数据
                insert_data = zip(cell_data_list, url_list, line_num_list)
                insert_data.sort(key=lambda x: x[0][0].upper())

                cell_data_list, url_list, line_num_list = zip(*insert_data)
                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                # 插入数据第一列到第三列
                nu = 3
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw.write(nu, predict_insert_location + 5, first_insert_data[line])
                        self.worksheet_sw.write_url(nu, predict_insert_location + 6, url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.write(nu, predict_insert_location + 7, second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw.write_url(nu, predict_insert_location + 8, url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw.write(nu, predict_insert_location + 8, cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw.merge_range(nu, predict_insert_location + 5, nu + line_num_list[line] - 1,
                                                      predict_insert_location + 5, first_insert_data[line], self.title_format)
                        self.worksheet_sw.merge_range(nu, predict_insert_location + 6, nu + line_num_list[line] - 1, predict_insert_location + 6, '')
                        self.worksheet_sw.write_url(nu, predict_insert_location + 6, url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.merge_range(nu, predict_insert_location + 7, nu + line_num_list[line] - 1, predict_insert_location + 7,
                                                      second_insert_data[line], self.title_format)
                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw.write_url(i, predict_insert_location + 8, url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw.write(line_num_list[line] + nu - 1, predict_insert_location + 8, cell_data_list[line][3:][-1])
                        nu += line_num_list[line]
            except:
                self.logger.print_message('predict_insert_SW_data fail url [ %s ]' % self._predict_url, self.__file_name)

    def predict_insert_SW_Original_data(self):
        if self.predict_insert_flag and self._predict_url:
            predict_insert_location = self.calculate_head_num(9, 0, 0, True)
            self.logger.print_message('SW Original predict_insert_location:\t%s' % predict_insert_location, self.__file_name)
            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self._predict_url, self.logger)
                if not is_validity:
                    return
            try:
                Silver_BkC_string, header_length, url_list, header_list, cell_data_list = self.predict_extract_object.predict_get_sw_data()
                self.worksheet_sw_original.write_rich_string(1, predict_insert_location + 1, self.red, self.save_miss_insert_bkc_string, self.header_format)
                self.worksheet_sw_original.write_rich_string(1, predict_insert_location + 2, self.red, Silver_BkC_string, self.header_format)

                if not header_list and not cell_data_list:
                    return

                # Set up some formats to use.
                self.worksheet_sw_original.conditional_format(3, predict_insert_location, 40, predict_insert_location + 1,
                                                     {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_sw_original.conditional_format(3, predict_insert_location + 2, 40, predict_insert_location + 2,
                                                     {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_sw_original.write_row(2, predict_insert_location + 5, header_list, self.header_format)
                # 插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)

                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                # 处理并插入数据
                insert_data = zip(cell_data_list, url_list, line_num_list)
                insert_data.sort(key=lambda x: x[0][0].upper())

                cell_data_list, url_list, line_num_list = zip(*insert_data)
                # print 'line_num_list:\t', line_num_list, len(line_num_list)

                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                # 插入数据第一列到第三列
                nu = 3
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw_original.write(nu, predict_insert_location + 5, first_insert_data[line])
                        self.worksheet_sw_original.write_url(nu, predict_insert_location + 6, url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_original.write(nu, predict_insert_location + 7, second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw_original.write_url(nu, predict_insert_location + 8, url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw_original.write(nu, predict_insert_location + 8, cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw_original.merge_range(nu, predict_insert_location + 5, nu + line_num_list[line] - 1,
                                                      predict_insert_location + 5, first_insert_data[line], self.title_format)
                        self.worksheet_sw_original.merge_range(nu, predict_insert_location + 6, nu + line_num_list[line] - 1, predict_insert_location + 6, '')
                        self.worksheet_sw_original.write_url(nu, predict_insert_location + 6, url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_original.merge_range(nu, predict_insert_location + 7, nu + line_num_list[line] - 1, predict_insert_location + 7,
                                                      second_insert_data[line], self.title_format)
                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw_original.write_url(i, predict_insert_location + 8, url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw_original.write(line_num_list[line] + nu - 1, predict_insert_location + 8, cell_data_list[line][3:][-1])
                        nu += line_num_list[line]
            except:
                self.logger.print_message('predict_insert_SW_data fail url [ %s ]' % self._predict_url, self.__file_name)

    def predict_insert_IFWI_Original_data(self):
        if self.predict_insert_flag and self._predict_url:
            predict_insert_location = self.calculate_head_num(6, 0, 0, True)
            self.logger.print_message('IFWI_Original predict_insert_location:\t%s' % predict_insert_location, self.__file_name)

            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self._predict_url, self.logger)
                if not is_validity:
                    return
            try:
                Silver_BkC_string, header_list, cell_data_list = self.predict_extract_object.predict_get_ifwi_data()
                self.worksheet_ifwi_original.conditional_format(4, predict_insert_location, 50, predict_insert_location + 1,
                                                               {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_ifwi_original.conditional_format(4, predict_insert_location + 2, 49, predict_insert_location + 2,
                                                               {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_ifwi_original.write_rich_string(1, predict_insert_location + 1, self.red, self.save_miss_insert_bkc_string, self.header_format)
                self.worksheet_ifwi_original.write_rich_string(1, predict_insert_location + 2, self.red, Silver_BkC_string, self.format1)

                if not header_list and not cell_data_list:
                    return

                self.worksheet_ifwi_original.write_row(3, predict_insert_location + 3, header_list, self.header_format)
                # 插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    # 以数字开头的元素前面加Nic
                    match_obj = re.match('\s+\d+', str(cell_data_list[ele][0]))
                    match_obj_back = re.match('^\d', str(cell_data_list[ele][0]))
                    if match_obj or match_obj_back:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                for begin in range(len(cell_data_list)):
                    self.worksheet_ifwi_original.write_row(begin + 4, predict_insert_location + 3, cell_data_list[begin])
            except:
                self.logger.print_message('predict_insert_IFWI_Original_data fail url [ %s ]' % self._predict_url, self.__file_name)

    def predict_insert_IFWI_data(self):
        if self.predict_insert_flag and self._predict_url:
            predict_insert_location = self.calculate_head_num(6, 0, 0, True)
            self.logger.print_message('IFWI predict_insert_location:\t%s' % predict_insert_location, self.__file_name)

            if self.on_off_line_save_flag == 'online':
                # 验证url有效性
                is_validity = verify_validity_url(self._predict_url, self.logger)
                if not is_validity:
                    return
            try:
                Silver_BkC_string, header_list, cell_data_list = self.predict_extract_object.predict_get_ifwi_data()
                self.worksheet_ifwi.conditional_format(4, predict_insert_location, 50, predict_insert_location + 1,
                                                               {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                # 标记0为黄色
                self.worksheet_ifwi.conditional_format(4, predict_insert_location + 2, 49, predict_insert_location + 2,
                                                               {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.yellow_data_format})
                self.worksheet_ifwi.write_rich_string(1, predict_insert_location + 1, self.red, self.save_miss_insert_bkc_string, self.header_format)
                self.worksheet_ifwi.write_rich_string(1, predict_insert_location + 2, self.red, Silver_BkC_string, self.format1)

                if not header_list and not cell_data_list:
                    return

                self.worksheet_ifwi.write_row(3, predict_insert_location + 3, header_list, self.header_format)
                # 插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    # 以数字开头的元素前面加Nic
                    match_obj = re.match('\s+\d+', str(cell_data_list[ele][0]))
                    match_obj_back = re.match('^\d', str(cell_data_list[ele][0]))
                    if match_obj or match_obj_back:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                cell_data_list.sort(key=lambda x: x[0].upper())
                for begin in range(len(cell_data_list)):
                    self.worksheet_ifwi.write_row(begin + 4, predict_insert_location + 3, cell_data_list[begin])
            except:
                self.logger.print_message('predict_insert_IFWI_data fail url [ %s ]' % self._predict_url, self.__file_name)

    def predict_insert_Platform_data(self):
        predict_insert_location = self.calculate_head_num(12, 0, 0, True)
        self.worksheet_platform.write_rich_string(1, predict_insert_location + 1, self.red, 'Candidate', self.format1)
        self.worksheet_platform.merge_range(2, predict_insert_location, 6, predict_insert_location, '', self.title_format)
        self.worksheet_platform.write_rich_string(2, predict_insert_location, self.red, self.save_miss_insert_bkc_string, self.title_format)

    def predict_insert_Mapping(self):
        reference_position_index = self.__WEEK_NUM + 6 - len(self.Silver_url_list) - 1
        self.worksheet_mapping.write(self.__WEEK_NUM + 6, reference_position_index - 1, self.save_miss_insert_bkc_string, self.header_format)
        self.worksheet_mapping.write(reference_position_index, self.__WEEK_NUM + 6, self.save_miss_insert_bkc_string, self.header_format)

    def predict_insert_CaseResult(self):
        predict_insert_location = self.calculate_head_num(41, 0, 0, True)
        self.worksheet_caseresult.write(6, predict_insert_location + 1 + 11 + 1, 'Candidate', self.format1)
        self.worksheet_caseresult.write(2, predict_insert_location + 11, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_save_miss_data(self):
        Silver_url_list = self.Silver_url_list
        # TODO 当自定义覆盖相应周数据时 需要处理日期对齐 start
        if self.keep_continuous == 'YES':
            Silver_url_list = self.section_Silver_url_list

        # TODO 新插入一个bkc字符串
        self.worksheet_save_miss.write(2, self.__WEEK_NUM + 2 - len(Silver_url_list) - 1, self.save_miss_insert_bkc_string, self.format1)

    def predict_insert_trend_data(self):
        pass



if __name__ == '__main__':
    conf = MachineConfig(r'C:\Users\pengzh5x\Desktop\machine_scripts\machineConfig\machine.conf')
    purl_bak_string = conf.get_node_info('real-time_control_parameter_value', 'default_purl_bak_string')
    Silver_url_list = []
    with open(r'C:\Users\pengzh5x\Desktop\machine_scripts\report_html\url_info.txt', 'r') as f:
        for line in f:
            if 'Silver' in line and purl_bak_string in line:
                Silver_url_list.append(line.strip('\n'))
        Silver_url_list = Silver_url_list[:50]
    print Silver_url_list
    print purl_bak_string
    start = time.time()
    cache = DiskCache(purl_bak_string)
    from machine_scripts.custom_log import WorkLogger

    log_time = time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time()))
    link_WW_week_string = 'Default'
    if Silver_url_list:
        WW_week_info_string = re.split('\D+', Silver_url_list[0].split('/')[-2])
        link_WW_week_string = WW_week_info_string[0] + 'WW' + WW_week_info_string[-1]

    _logger = WorkLogger(log_filename='machine.log', log_level=20, create_log_flag= True, log_time=log_time)
    obj = InsertDataIntoExcel(verify_flag=True, purl_bak_string=purl_bak_string,
                                        link_WW_week_string=link_WW_week_string, cache=cache,
                                        silver_url_list=Silver_url_list,
                                        section_Silver_url_list=[], logger=_logger,
                                        log_time=log_time, keep_continuous=False)
    func_name_list = obj.return_name().keys()
    call_func_list = [func for func in func_name_list if func.startswith('insert')]
    if 'insert_CaseResult' in call_func_list:
        call_func_list.remove('insert_CaseResult')
        call_func_list.append('insert_CaseResult')
        print 'call_func_list:\t%s' % call_func_list

    for func in call_func_list:
        getattr(obj, func)(1)
    obj.predict_insert_IFWI_data()
    obj.predict_insert_IFWI_Original_data()
    obj.predict_insert_SW_data()
    obj.predict_insert_SW_Original_data()
    obj.close_workbook()
    _logger.file_close()

    print time.time() - start